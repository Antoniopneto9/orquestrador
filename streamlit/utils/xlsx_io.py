"""
xlsx_io.py — Read/write interface for the intentional_season.xlsx workbook.

All functions handle a missing file gracefully by calling ensure_workbook()
before performing any operation.
"""

from pathlib import Path
from datetime import date, timedelta
import calendar

import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook

# Import config values
from utils.config import (
    XLSX_PATH,
    LOGS_BASE_PATH,
    HABITS_BY_MONTH,
    ALL_COMPLAINTS,
    MONTHS_RANGE,
    GOAL_CATEGORIES,
    EXPENSE_CATEGORIES,
    LEITURA_SHEET,
    LEITURA_COLS,
    SHEET_RASTREADOR,
    SHEET_QUEIXAS,
    SHEET_REVISAO,
    SHEET_LINHA_TEMPO,
    SHEET_DASHBOARD,
    SHEET_FINANCEIRO,
    MENSAL_SHEETS,
    SEASON_START,
    SEASON_END,
    mensal_sheet_name,
)

# ── Constants ─────────────────────────────────────────────────────────────────

# All 19 habits (max list from August onward)
ALL_HABITS: list[str] = HABITS_BY_MONTH["2026-08"]

# Portuguese weekday abbreviations (Monday=0)
WEEKDAY_PT = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]

# Season days: list of date objects from start to end inclusive
def _season_dates() -> list[date]:
    days = []
    current = SEASON_START
    while current <= SEASON_END:
        days.append(current)
        current += timedelta(days=1)
    return days

SEASON_DATES = _season_dates()


# ── Workbook initialization ───────────────────────────────────────────────────

def initialize_workbook() -> None:
    """Create the .xlsx file with all sheets and pre-filled row skeletons."""
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    _create_rastreador(wb)
    _create_queixas(wb)
    _create_revisao_semanal(wb)
    _create_linha_do_tempo(wb)
    _create_dashboard(wb)
    _create_financeiro(wb)
    _create_leitura(wb)
    _create_mensal_sheets(wb)

    wb.save(XLSX_PATH)


def _create_rastreador(wb: Workbook) -> None:
    """365-row tracker sheet: one row per season day, one column per habit."""
    ws = wb.create_sheet(SHEET_RASTREADOR)

    # Header row
    headers = ["dia", "data", "dia_semana"] + ALL_HABITS + ["total", "percentual"]
    ws.append(headers)

    # One row per day
    for i, d in enumerate(SEASON_DATES, start=1):
        weekday_name = WEEKDAY_PT[d.weekday()]
        # Empty strings for all habit columns, total, percentual
        row = [i, d.strftime("%Y-%m-%d"), weekday_name] + [""] * len(ALL_HABITS) + [0, 0.0]
        ws.append(row)


def _create_queixas(wb: Workbook) -> None:
    """365-row complaints sheet: one row per season day, one column per complaint."""
    ws = wb.create_sheet(SHEET_QUEIXAS)

    headers = ["dia", "data"] + ALL_COMPLAINTS + ["media", "observacao"]
    ws.append(headers)

    for i, d in enumerate(SEASON_DATES, start=1):
        row = [i, d.strftime("%Y-%m-%d")] + [0] * len(ALL_COMPLAINTS) + [0.0, ""]
        ws.append(row)


def _create_revisao_semanal(wb: Workbook) -> None:
    """52-row weekly review sheet."""
    ws = wb.create_sheet(SHEET_REVISAO)

    headers = ["semana", "data_inicio", "funcionou", "quebrou", "ajuste", "nota"]
    ws.append(headers)

    # One row per week, starting from SEASON_START
    for week_num in range(1, 53):
        week_start = SEASON_START + timedelta(weeks=week_num - 1)
        row = [week_num, week_start.strftime("%Y-%m-%d"), "", "", "", ""]
        ws.append(row)


def _create_linha_do_tempo(wb: Workbook) -> None:
    """54-row goal timeline sheet: one row per goal, columns for each month."""
    ws = wb.create_sheet(SHEET_LINHA_TEMPO)

    headers = ["meta", "categoria"] + MONTHS_RANGE + ["mes_previsto", "observacao"]
    ws.append(headers)

    # Pre-fill 54 placeholder goal rows
    for i in range(1, 55):
        row = [f"Meta {i}", GOAL_CATEGORIES[0]] + [""] * len(MONTHS_RANGE) + ["", ""]
        ws.append(row)


def _create_dashboard(wb: Workbook) -> None:
    """8-row dashboard: one row per base habit (March list)."""
    ws = wb.create_sheet(SHEET_DASHBOARD)

    headers = ["habito", "cumpridos", "falhos", "sem_resposta", "percentual", "observacao"]
    ws.append(headers)

    # Only the 8 base habits (March list)
    base_habits = HABITS_BY_MONTH["2026-03"]
    for habit in base_habits:
        row = [habit, 0, 0, 0, 0.0, ""]
        ws.append(row)


def _create_financeiro(wb: Workbook) -> None:
    """365-row financial sheet: one row per season day."""
    ws = wb.create_sheet(SHEET_FINANCEIRO)

    headers = ["dia", "data", "descricao", "valor", "categoria", "divida_total", "patrimonio"]
    ws.append(headers)

    for i, d in enumerate(SEASON_DATES, start=1):
        row = [i, d.strftime("%Y-%m-%d"), "", 0.0, "", 0.0, 0.0]
        ws.append(row)


def _create_mensal_sheets(wb: Workbook) -> None:
    """One sheet per month (Mar/2026 to Feb/2027) with daily rows."""
    # Month ranges: (year, month) pairs
    months = [(2026, m) for m in range(3, 13)] + [(2027, m) for m in range(1, 3)]

    for year, month in months:
        sheet_name = mensal_sheet_name(year, month)
        ws = wb.create_sheet(sheet_name)

        headers = ["dia", "data", "dia_semana"] + ALL_HABITS + ["total", "percentual"]
        ws.append(headers)

        # Days in this month that fall within the season
        days_in_month = calendar.monthrange(year, month)[1]
        day_num = 1  # row number within this month sheet
        for day in range(1, days_in_month + 1):
            d = date(year, month, day)
            if d < SEASON_START or d > SEASON_END:
                continue
            weekday_name = WEEKDAY_PT[d.weekday()]
            row = [day_num, d.strftime("%Y-%m-%d"), weekday_name] + [""] * len(ALL_HABITS) + [0, 0.0]
            ws.append(row)
            day_num += 1


def ensure_workbook() -> None:
    """Create the workbook if it does not exist yet."""
    if not XLSX_PATH.exists():
        initialize_workbook()


# Run on import so every function below can assume the file exists
ensure_workbook()


# ── Read helpers ──────────────────────────────────────────────────────────────

def load_sheet(sheet_name: str) -> pd.DataFrame:
    """Load a sheet as a DataFrame. Returns empty DataFrame if sheet missing."""
    ensure_workbook()
    try:
        df = pd.read_excel(XLSX_PATH, sheet_name=sheet_name, engine="openpyxl")
        return df
    except Exception:
        return pd.DataFrame()


def get_row_by_date(sheet_name: str, date_str: str) -> dict:
    """Return the row matching date_str ('YYYY-MM-DD') as a dict. Empty dict if not found."""
    df = load_sheet(sheet_name)
    if df.empty or "data" not in df.columns:
        return {}

    # Normalize the data column to string for comparison
    df["data"] = df["data"].astype(str).str[:10]
    match = df[df["data"] == date_str]
    if match.empty:
        return {}
    return match.iloc[0].to_dict()


# ── Write helpers ─────────────────────────────────────────────────────────────

def save_cell(sheet_name: str, date_str: str, column_name: str, value) -> None:
    """
    Find the row in sheet_name where data == date_str,
    then write value into column_name.
    """
    ensure_workbook()
    wb = load_workbook(XLSX_PATH)

    if sheet_name not in wb.sheetnames:
        wb.close()
        return

    ws = wb[sheet_name]

    # Find column index (1-based) by header name
    headers = [cell.value for cell in ws[1]]
    if column_name not in headers:
        wb.close()
        return
    col_idx = headers.index(column_name) + 1

    # Find row index by matching data column
    data_col_idx = headers.index("data") + 1 if "data" in headers else None
    if data_col_idx is None:
        wb.close()
        return

    for row in ws.iter_rows(min_row=2):
        cell_value = row[data_col_idx - 1].value
        # Normalize: openpyxl may return datetime objects
        if hasattr(cell_value, "strftime"):
            cell_value = cell_value.strftime("%Y-%m-%d")
        if str(cell_value)[:10] == date_str:
            row[col_idx - 1].value = value
            break

    wb.save(XLSX_PATH)
    wb.close()


def _create_leitura(wb: Workbook) -> None:
    """Create empty leitura sheet with header only."""
    ws = wb.create_sheet(LEITURA_SHEET)
    ws.append(LEITURA_COLS)


def append_row(sheet_name: str, data_dict: dict, default_headers: list | None = None) -> None:
    """
    Append a new row to sheet_name. If sheet doesn't exist and default_headers
    is provided, creates the sheet first. Columns ordered by sheet header row.
    """
    ensure_workbook()
    wb = load_workbook(XLSX_PATH)

    if sheet_name not in wb.sheetnames:
        if default_headers is None:
            wb.close()
            return
        ws = wb.create_sheet(sheet_name)
        ws.append(default_headers)
    else:
        ws = wb[sheet_name]

    headers = [cell.value for cell in ws[1]]
    row_values = [data_dict.get(col, "") for col in headers]
    ws.append(row_values)

    wb.save(XLSX_PATH)
    wb.close()


def save_row(sheet_name: str, row_key_col: str, row_key_val, data_dict: dict) -> None:
    """
    Find the row where row_key_col == row_key_val, then update all columns
    in data_dict. Creates or overwrites only the matching row.
    """
    ensure_workbook()
    wb = load_workbook(XLSX_PATH)

    if sheet_name not in wb.sheetnames:
        wb.close()
        return

    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]

    if row_key_col not in headers:
        wb.close()
        return

    key_col_idx = headers.index(row_key_col) + 1

    for row in ws.iter_rows(min_row=2):
        cell_val = row[key_col_idx - 1].value
        # Normalize datetime cells
        if hasattr(cell_val, "strftime"):
            cell_val = cell_val.strftime("%Y-%m-%d")

        if str(cell_val) == str(row_key_val):
            # Write each column in data_dict
            for col_name, value in data_dict.items():
                if col_name in headers:
                    col_idx = headers.index(col_name) + 1
                    row[col_idx - 1].value = value
            break

    wb.save(XLSX_PATH)
    wb.close()


# ── Agent log helpers ─────────────────────────────────────────────────────────

def get_latest_log(agent_name: str) -> tuple[str, str]:
    """
    Return (filename, content) for the most recent log file in
    logs/sessoes/{agent_name}/. Returns ("", "") if no logs exist.
    """
    agent_dir = LOGS_BASE_PATH / agent_name

    if not agent_dir.exists():
        return ("", "")

    # Gather all files sorted by modification time (newest first)
    log_files = sorted(
        agent_dir.iterdir(),
        key=lambda f: f.stat().st_mtime,
        reverse=True,
    )
    # Filter to actual files (not directories)
    log_files = [f for f in log_files if f.is_file()]

    if not log_files:
        return ("", "")

    latest = log_files[0]
    content = latest.read_text(encoding="utf-8", errors="replace")
    return (latest.name, content)


def list_agent_logs(agent_name: str) -> list[str]:
    """Return sorted list of log filenames for the given agent (newest first)."""
    agent_dir = LOGS_BASE_PATH / agent_name
    if not agent_dir.exists():
        return []
    files = sorted(
        [f.name for f in agent_dir.iterdir() if f.is_file()],
        reverse=True,
    )
    return files
